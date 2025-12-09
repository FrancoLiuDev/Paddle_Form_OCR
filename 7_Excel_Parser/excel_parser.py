#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 檔案解析與分析工具
分析 Excel 檔案結構、內容和統計資訊
"""

import sys
import json
import os
from typing import Dict, List, Any, Optional
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime


class ExcelParser:
    """Excel 檔案解析器"""
    
    def __init__(self, file_path: str):
        """初始化解析器
        
        Args:
            file_path: Excel 檔案路徑
        """
        self.file_path = file_path
        self.workbook = None
        
    def parse(self) -> Dict[str, Any]:
        """解析 Excel 檔案
        
        Returns:
            解析結果字典
        """
        try:
            # 載入工作簿
            self.workbook = openpyxl.load_workbook(
                self.file_path, 
                data_only=True,
                read_only=False
            )
            
            # 基本資訊
            file_stat = os.stat(self.file_path)
            
            result = {
                "success": True,
                "file_info": {
                    "path": self.file_path,
                    "name": os.path.basename(self.file_path),
                    "size": file_stat.st_size,
                    "size_formatted": self._format_size(file_stat.st_size),
                    "modified": datetime.fromtimestamp(file_stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                },
                "workbook_info": {
                    "total_sheets": len(self.workbook.sheetnames),
                    "sheet_names": self.workbook.sheetnames,
                    "active_sheet": self.workbook.active.title if self.workbook.active else None
                },
                "sheets": []
            }
            
            # 分析每個工作表
            for sheet_name in self.workbook.sheetnames:
                sheet_info = self._parse_sheet(sheet_name)
                result["sheets"].append(sheet_info)
            
            return result
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "error_type": type(e).__name__,
                "file_path": self.file_path
            }
        finally:
            if self.workbook:
                self.workbook.close()
    
    def _parse_sheet(self, sheet_name: str) -> Dict[str, Any]:
        """解析單個工作表
        
        Args:
            sheet_name: 工作表名稱
            
        Returns:
            工作表解析結果
        """
        sheet = self.workbook[sheet_name]
        
        # 取得使用範圍
        max_row = sheet.max_row
        max_col = sheet.max_column
        min_row = sheet.min_row
        min_col = sheet.min_column
        
        # 讀取所有資料（限制最多 1000 行避免記憶體問題）
        data_rows = []
        read_limit = min(max_row, 1000)
        
        for row_idx in range(min_row, read_limit + 1):
            row_data = []
            for col_idx in range(min_col, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell_info = {
                    "value": self._serialize_value(cell.value),
                    "type": self._get_cell_type(cell),
                    "coordinate": cell.coordinate,
                    "formatted": str(cell.value) if cell.value is not None else ""
                }
                
                # 如果有公式
                if cell.data_type == 'f':
                    cell_info["formula"] = cell.value
                
                # 如果有超連結
                if cell.hyperlink:
                    cell_info["hyperlink"] = cell.hyperlink.target
                
                row_data.append(cell_info)
            data_rows.append(row_data)
        
        # 檢測標題行（假設第一行是標題）
        headers = []
        if max_row > 0:
            for col_idx in range(min_col, max_col + 1):
                cell_value = sheet.cell(row=min_row, column=col_idx).value
                headers.append(str(cell_value) if cell_value is not None else f"Column_{col_idx}")
        
        # 統計資訊
        stats = self._calculate_statistics(sheet, min_row, max_row, min_col, max_col)
        
        # 合併儲存格資訊
        merged_cells = [str(merged) for merged in sheet.merged_cells.ranges]
        
        return {
            "name": sheet_name,
            "range": {
                "min_row": min_row,
                "max_row": max_row,
                "min_column": min_col,
                "max_column": max_col,
                "total_rows": max_row - min_row + 1,
                "total_columns": max_col - min_col + 1
            },
            "headers": headers,
            "data": data_rows,
            "statistics": stats,
            "merged_cells": merged_cells,
            "has_more": max_row > read_limit
        }
    
    def _get_cell_type(self, cell) -> str:
        """取得儲存格資料類型
        
        Args:
            cell: 儲存格物件
            
        Returns:
            類型字串
        """
        if cell.value is None:
            return "empty"
        elif isinstance(cell.value, bool):
            return "boolean"
        elif isinstance(cell.value, (int, float)):
            return "number"
        elif isinstance(cell.value, str):
            return "string"
        elif isinstance(cell.value, datetime):
            return "datetime"
        else:
            return "unknown"
    
    def _serialize_value(self, value: Any) -> Any:
        """序列化儲存格值
        
        Args:
            value: 儲存格值
            
        Returns:
            可序列化的值
        """
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        return value
    
    def _calculate_statistics(self, sheet, min_row: int, max_row: int, 
                              min_col: int, max_col: int) -> Dict[str, Any]:
        """計算工作表統計資訊
        
        Args:
            sheet: 工作表物件
            min_row: 最小行數
            max_row: 最大行數
            min_col: 最小列數
            max_col: 最大列數
            
        Returns:
            統計資訊字典
        """
        total_cells = (max_row - min_row + 1) * (max_col - min_col + 1)
        empty_cells = 0
        numeric_cells = 0
        text_cells = 0
        formula_cells = 0
        date_cells = 0
        boolean_cells = 0
        
        numeric_values = []
        
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, 
                                   min_col=min_col, max_col=max_col):
            for cell in row:
                if cell.value is None:
                    empty_cells += 1
                elif isinstance(cell.value, bool):
                    boolean_cells += 1
                elif isinstance(cell.value, (int, float)):
                    numeric_cells += 1
                    numeric_values.append(float(cell.value))
                elif isinstance(cell.value, str):
                    text_cells += 1
                elif isinstance(cell.value, datetime):
                    date_cells += 1
                
                if cell.data_type == 'f':
                    formula_cells += 1
        
        # 數值統計
        numeric_stats = {}
        if numeric_values:
            numeric_stats = {
                "count": len(numeric_values),
                "sum": sum(numeric_values),
                "average": sum(numeric_values) / len(numeric_values),
                "min": min(numeric_values),
                "max": max(numeric_values)
            }
        
        return {
            "total_cells": total_cells,
            "empty_cells": empty_cells,
            "numeric_cells": numeric_cells,
            "text_cells": text_cells,
            "formula_cells": formula_cells,
            "date_cells": date_cells,
            "boolean_cells": boolean_cells,
            "fill_rate": round((total_cells - empty_cells) / total_cells * 100, 2) if total_cells > 0 else 0,
            "numeric_stats": numeric_stats
        }
    
    def _format_size(self, size_bytes: int) -> str:
        """格式化檔案大小
        
        Args:
            size_bytes: 位元組大小
            
        Returns:
            格式化字串
        """
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size_bytes < 1024.0:
                return f"{size_bytes:.2f} {unit}"
            size_bytes /= 1024.0
        return f"{size_bytes:.2f} TB"


def main():
    """主程式"""
    if len(sys.argv) < 2:
        print(json.dumps({
            "success": False,
            "error": "請提供 Excel 檔案路徑",
            "usage": "python excel_parser.py <file_path>"
        }, ensure_ascii=False))
        sys.exit(1)
    
    file_path = sys.argv[1]
    
    if not os.path.exists(file_path):
        print(json.dumps({
            "success": False,
            "error": f"檔案不存在: {file_path}"
        }, ensure_ascii=False))
        sys.exit(1)
    
    parser = ExcelParser(file_path)
    result = parser.parse()
    
    # 輸出 JSON 結果
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
