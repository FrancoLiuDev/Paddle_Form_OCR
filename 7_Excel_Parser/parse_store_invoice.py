#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
解析「門市請款明細表」工作表
將每個門市編碼下的項次列出來
"""

import os
import sys
import json
from typing import Dict, List, Any
import openpyxl


class StoreInvoiceParser:
    """門市請款明細表解析器"""
    
    def __init__(self, file_path: str):
        """初始化解析器
        
        Args:
            file_path: Excel 檔案路徑
        """
        self.file_path = file_path
        self.workbook = None
        
    def parse(self) -> Dict[str, Any]:
        """解析門市請款明細表
        
        Returns:
            解析結果字典
        """
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            
            # 找到「門市請款明細表」工作表
            sheet_name = "門市請款明細表"
            if sheet_name not in self.workbook.sheetnames:
                return {
                    "success": False,
                    "error": f"找不到工作表: {sheet_name}",
                    "available_sheets": self.workbook.sheetnames
                }
            
            sheet = self.workbook[sheet_name]
            
            # 讀取標題行（第1行）
            headers = []
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col_idx).value
                headers.append(str(cell_value) if cell_value else f"Column_{col_idx}")
            
            # 找出關鍵欄位的索引
            col_indices = self._find_column_indices(headers)
            
            # 解析資料
            stores_data = self._parse_store_data(sheet, col_indices)
            
            result = {
                "success": True,
                "sheet_name": sheet_name,
                "total_rows": sheet.max_row,
                "headers": headers,
                "column_mapping": col_indices,
                "total_stores": len(stores_data),
                "stores": stores_data
            }
            
            return result
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "error_type": type(e).__name__
            }
        finally:
            if self.workbook:
                self.workbook.close()
    
    def _find_column_indices(self, headers: List[str]) -> Dict[str, int]:
        """找出所有欄位的索引（動態取得）
        
        Args:
            headers: 標題列表
            
        Returns:
            欄位名稱到索引的映射
        """
        indices = {}
        
        # 動態讀取所有欄位，使用清理後的標題作為欄位名
        for col_idx, header in enumerate(headers, 1):
            if header and header != f"Column_{col_idx}":
                # 清理標題名稱（移除換行符等）
                field_name = header.strip().replace('\n', '')
                indices[field_name] = col_idx
            else:
                # 對於空標題，使用 Column_N 命名
                indices[f"Column_{col_idx}"] = col_idx
        
        return indices
    
    def _parse_store_data(self, sheet, col_indices: Dict[str, int]) -> List[Dict[str, Any]]:
        """解析門市資料
        
        Args:
            sheet: 工作表物件
            col_indices: 欄位索引映射
            
        Returns:
            門市資料列表
        """
        stores_data = []
        current_store = None
        
        # 動態查找關鍵欄位
        store_code_col = None
        store_name_col = None
        item_no_col = None
        
        for field_name, col_idx in col_indices.items():
            field_clean = field_name.replace('\n', '').replace(' ', '')
            if '門市編碼' in field_clean or field_clean == '門市編碼':
                store_code_col = col_idx
            elif field_clean == '門市':
                store_name_col = col_idx
            elif field_clean == '項次':
                item_no_col = col_idx
        
        # 從第2行開始讀取（第1行是標題）
        for row_idx in range(2, sheet.max_row + 1):
            # 讀取門市編碼
            if store_code_col:
                store_code = sheet.cell(row=row_idx, column=store_code_col).value
                
                # 如果有新的門市編碼，創建新的門市記錄
                if store_code and str(store_code).strip():
                    # 保存上一個門市
                    if current_store:
                        stores_data.append(current_store)
                    
                    # 創建新門市記錄
                    current_store = {
                        "門市編碼": str(store_code).strip(),
                        "項次列表": []
                    }
                    
                    # 讀取門市名稱
                    if store_name_col:
                        store_name = sheet.cell(row=row_idx, column=store_name_col).value
                        if store_name:
                            current_store["門市名稱"] = str(store_name).strip()
                        else:
                            current_store["門市名稱"] = ""
                    else:
                        current_store["門市名稱"] = ""
            
            # 讀取項次資料
            if current_store:
                item_data = self._read_item_data(sheet, row_idx, col_indices, item_no_col)
                if item_data:
                    current_store["項次列表"].append(item_data)
        
        # 保存最後一個門市
        if current_store:
            stores_data.append(current_store)
        
        return stores_data
    
    def _read_item_data(self, sheet, row_idx: int, col_indices: Dict[str, int], item_no_col: int = None) -> Dict[str, Any]:
        """讀取單筆項次資料（動態讀取所有欄位）
        
        Args:
            sheet: 工作表物件
            row_idx: 行索引
            col_indices: 欄位索引映射
            item_no_col: 項次欄位索引
            
        Returns:
            項次資料字典
        """
        item = {}
        
        # 讀取所有欄位
        for field_name, col_idx in col_indices.items():
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            
            # 格式化值
            if cell_value is None:
                item[field_name] = None
            elif isinstance(cell_value, (int, float)):
                item[field_name] = cell_value
            else:
                item[field_name] = str(cell_value).strip()
        
        # 檢查是否有項次值（如果找到項次欄位）
        if item_no_col:
            if sheet.cell(row=row_idx, column=item_no_col).value:
                return item
        else:
            # 如果沒有明確的項次欄位，檢查是否有任何非空值
            has_data = any(v is not None and v != '' for v in item.values())
            if has_data:
                return item
        
        return None


def main():
    # 目標檔案（檢查多個可能的位置）
    possible_paths = [
        "input/example.xlsx",
        "../6_Desktop_App/input/example.xlsx",
    ]
    
    example_file = None
    for path in possible_paths:
        if os.path.exists(path):
            example_file = path
            break
    
    if not example_file:
        print(f"❌ 找不到檔案，嘗試過的路徑:")
        for path in possible_paths:
            print(f"   - {path}")
        sys.exit(1)
    
    print(f"📂 正在解析門市請款明細表...")
    print(f"   檔案: {example_file}")
    
    # 執行解析
    parser = StoreInvoiceParser(example_file)
    result = parser.parse()
    
    if not result.get("success"):
        print(f"❌ 解析失敗: {result.get('error')}")
        sys.exit(1)
    
    # 確保 result 目錄存在
    result_dir = "result"
    os.makedirs(result_dir, exist_ok=True)
    
    # 輸出完整 JSON
    output_file = os.path.join(result_dir, "store_invoice_detail.json")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    
    print(f"✅ 結果已輸出: {output_file}")
    print(f"   檔案大小: {os.path.getsize(output_file):,} bytes")
    
    # 顯示統計資訊
    print(f"\n📊 解析統計:")
    print(f"   總行數: {result['total_rows']:,}")
    print(f"   總欄位數: {len(result['headers'])}")
    print(f"   門市數量: {result['total_stores']}")
    
    # 顯示欄位列表
    print(f"\n📋 欄位列表:")
    for i, (field_name, col_idx) in enumerate(result['column_mapping'].items(), 1):
        print(f"   {i}. {field_name} (第{col_idx}欄)")
        if i >= 10:
            remaining = len(result['column_mapping']) - 10
            if remaining > 0:
                print(f"   ... 還有 {remaining} 個欄位")
            break
    
    # 顯示前幾個門市的資訊
    print(f"\n📍 門市資訊 (前5個):")
    for i, store in enumerate(result['stores'][:5], 1):
        print(f"   {i}. {store['門市編碼']} - {store.get('門市名稱', '')}")
        print(f"      項次數量: {len(store['項次列表'])}")
        if store['項次列表']:
            # 顯示第一筆項次的關鍵資訊
            first_item = store['項次列表'][0]
            item_no = first_item.get('項次', first_item.get('項次', 'N/A'))
            model = first_item.get('機型', 'N/A')
            print(f"      第一筆: 項次={item_no}, 機型={model}")
    
    if result['total_stores'] > 5:
        print(f"   ... 還有 {result['total_stores'] - 5} 個門市")
    
    print(f"\n✨ 完成！")


if __name__ == "__main__":
    main()
